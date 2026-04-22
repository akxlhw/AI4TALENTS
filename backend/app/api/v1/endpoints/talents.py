"""
Talents API endpoints.
Provides talent list, detail, and filtering.

Architecture: Endpoint -> Service -> Repository
"""
from __future__ import annotations

import asyncio
import csv
import io

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_async_session
from app.schemas.common import PaginatedResponse
from app.schemas.overview import SelectedWorkResponse, TalentDetail, TalentSummary, TechTagItem
from app.services.talent_service import TalentService

router = APIRouter(prefix="/talents", tags=["Talents"])


@router.get(
    "",
    response_model=PaginatedResponse[TalentSummary],
    summary="获取人才列表",
    description="分页查询人才列表，支持多种筛选条件",
)
async def list_talents(
    school_id: int | None = Query(None, description="按学校ID筛选"),
    country_code: str | None = Query(None, description="按国家代码筛选 (如 US, CN)"),
    role_type: str | None = Query(None, description="按角色类型筛选 (professor/student/graduated/unknown)"),
    min_works: int | None = Query(None, description="最小论文数"),
    min_citations: int | None = Query(None, description="最小引用数"),
    keyword: str | None = Query(None, description="搜索关键词"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Get paginated list of talents.

    Supports filtering by:
    - school_id: Filter by school
    - country_code: Filter by country code (via school)
    - role_type: Filter by role type
    - min_works: Minimum works count
    - min_citations: Minimum citation count
    - keyword: Search in name, English name, and title
    """
    service = TalentService(session)
    talents, total = await service.get_talent_list(
        school_id=school_id,
        country_code=country_code,
        role_type=role_type,
        min_works=min_works,
        min_citations=min_citations,
        keyword=keyword,
        page=page,
        page_size=page_size,
    )

    items = [
        TalentSummary(
            talent_id=talent.talent_id,
            name=talent.name,
            name_en=talent.name_en,
            orcid=talent.orcid,
            role_type=talent.role_type,
            role_confidence=talent.role_confidence,
            school_id=talent.school_id,
            school_name=talent.school.school_name if talent.school else None,
            current_title=talent.current_title,
            works_count=talent.works_count,
            cited_by_count=talent.cited_by_count,
            h_index=talent.h_index,
            topic_tags=talent.topic_tags or [],
            openalex_topics=talent.openalex_topics or [],
        )
        for talent in talents
    ]

    return PaginatedResponse.create(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get(
    "/{talent_id}",
    response_model=TalentDetail,
    summary="获取人才详情",
    description="返回人才的详细信息，包括研究兴趣、代表作品等",
)
async def get_talent(
    talent_id: int,
    session: AsyncSession = Depends(get_async_session),
):
    """
    Get detailed information about a specific talent.

    Returns:
    - Basic talent information
    - School affiliation
    - Research statistics
    - Role profile
    - Selected works
    - Tech tags
    """
    from sqlalchemy import select

    from app.models.tech_domain import TalentTechTag, TechDirection, TechDomain

    service = TalentService(session)
    talent = await service.get_talent_by_id(talent_id)

    if not talent:
        raise HTTPException(status_code=404, detail="Talent not found")

    # Build selected works list
    selected_works = [
        SelectedWorkResponse(
            work_id=work.work_id,
            title=work.title,
            publication_year=work.publication_year,
            venue_name=work.venue_name,
            citation_count=work.citation_count,
            doi=work.doi,
        )
        for work in (talent.selected_works or [])
    ]

    # Fetch tech tags
    tech_tags = []
    result = await session.execute(
        select(TalentTechTag, TechDomain, TechDirection)
        .join(TechDomain, TalentTechTag.tech_domain_id == TechDomain.tech_domain_id)
        .outerjoin(TechDirection, TalentTechTag.tech_direction_id == TechDirection.tech_direction_id)
        .where(TalentTechTag.talent_id == talent_id)
    )
    for _tag, domain, direction in result.fetchall():
        tech_tags.append(TechTagItem(
            tech_domain_id=domain.tech_domain_id,
            tech_domain_name=domain.domain_name,
            tech_direction_id=direction.tech_direction_id if direction else None,
            tech_direction_name=direction.direction_name if direction else None,
        ))

    return TalentDetail(
        talent_id=talent.talent_id,
        name=talent.name,
        name_en=talent.name_en,
        orcid=talent.orcid,
        role_type=talent.role_type,
        role_confidence=talent.role_confidence,
        school_id=talent.school_id,
        school_name=talent.school.school_name if talent.school else None,
        current_title=talent.current_title,
        works_count=talent.works_count,
        cited_by_count=talent.cited_by_count,
        h_index=talent.h_index,
        latest_active_year=talent.latest_active_year,
        topic_tags=talent.topic_tags or [],
        openalex_topics=talent.openalex_topics or [],
        tech_tags=tech_tags,
        summary=talent.summary,
        department_name=talent.department_name,
        lab_name=talent.lab_name,
        role_reason=talent.role_profile.role_reason if talent.role_profile else None,
        academic_age=talent.role_profile.academic_age if talent.role_profile else None,
        selected_works=selected_works,
    )


@router.get(
    "/{talent_id}/works",
    response_model=list[SelectedWorkResponse],
    summary="获取人才代表作品",
    description="返回人才的代表作品列表",
)
async def get_talent_works(
    talent_id: int,
    limit: int = Query(10, ge=1, le=50, description="返回数量限制"),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Get selected works for a specific talent.

    Returns list of representative works ordered by citation count.
    """
    service = TalentService(session)

    # Verify talent exists
    if not await service.talent_exists(talent_id):
        raise HTTPException(status_code=404, detail="Talent not found")

    works = await service.get_selected_works(talent_id, limit=limit)

    return [
        SelectedWorkResponse(
            work_id=work.work_id,
            title=work.title,
            publication_year=work.publication_year,
            venue_name=work.venue_name,
            citation_count=work.citation_count,
            doi=work.doi,
        )
        for work in works
    ]


@router.post(
    "/export",
    summary="导出候选人",
    description="导出选中的候选人数据为CSV或Excel格式",
)
async def export_talents(
    format: str = Query("csv", description="导出格式: csv 或 xlsx"),
    talent_ids: list[int] = Query(..., description="要导出的人才ID列表"),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Export selected talents to CSV or Excel format.
    """
    from openpyxl import Workbook

    service = TalentService(session)

    # Fetch talents by IDs
    talents = await service.get_talents_by_ids(talent_ids)

    if not talents:
        raise HTTPException(status_code=404, detail="未找到要导出的人才")

    # Prepare data
    headers = ["ID", "姓名", "英文名", "ORCID", "角色", "学校", "职位", "论文数", "引用数", "H指数", "研究方向"]
    rows = []
    for t in talents:
        rows.append([
            t.talent_id,
            t.name,
            t.name_en or "",
            t.orcid or "",
            t.role_type,
            t.school.school_name if t.school else "",
            t.current_title or "",
            t.works_count,
            t.cited_by_count,
            t.h_index,
            ", ".join(t.topic_tags or []),
        ])

    if format == "xlsx":
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "候选人导出"

        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=talents_export.xlsx"},
        )
    else:
        # Create CSV
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(rows)
        buffer.seek(0)

        return StreamingResponse(
            io.BytesIO(buffer.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=talents_export.csv"},
        )


@router.post(
    "/compare",
    summary="对比候选人",
    description="获取多个候选人的对比数据",
)
async def compare_talents(
    talent_ids: list[int] = Query(..., description="要对比的人才ID列表 (2-4个)"),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Compare multiple talents (2-4) side by side.
    Returns detailed comparison data.
    """
    if len(talent_ids) < 2 or len(talent_ids) > 4:
        raise HTTPException(status_code=400, detail="请选择2-4位候选人进行对比")

    service = TalentService(session)

    # Fetch talents by IDs
    talents = await service.get_talents_by_ids(talent_ids)

    if not talents:
        raise HTTPException(status_code=404, detail="未找到要对比的人才")

    # Build comparison data
    comparison_data = []
    for t in talents:
        comparison_data.append({
            "talent_id": t.talent_id,
            "name": t.name,
            "name_en": t.name_en,
            "orcid": t.orcid,
            "role_type": t.role_type,
            "school_id": t.school_id,
            "school_name": t.school.school_name if t.school else None,
            "current_title": t.current_title,
            "department_name": t.department_name,
            "lab_name": t.lab_name,
            "works_count": t.works_count,
            "cited_by_count": t.cited_by_count,
            "h_index": t.h_index,
            "latest_active_year": t.latest_active_year,
            "topic_tags": t.topic_tags or [],
            "openalex_topics": t.openalex_topics or [],
            "academic_age": t.role_profile.academic_age if t.role_profile else None,
        })

    return {
        "talents": comparison_data,
        "comparison_fields": [
            {"key": "name", "label": "姓名"},
            {"key": "role_type", "label": "角色"},
            {"key": "school_name", "label": "学校"},
            {"key": "current_title", "label": "职位"},
            {"key": "department_name", "label": "院系"},
            {"key": "works_count", "label": "论文数"},
            {"key": "cited_by_count", "label": "引用数"},
            {"key": "h_index", "label": "H指数"},
            {"key": "latest_active_year", "label": "最近活跃年份"},
            {"key": "academic_age", "label": "学术年龄"},
            {"key": "topic_tags", "label": "研究方向"},
        ]
    }


@router.post(
    "/collaborations/generate-sample",
    summary="生成示例合作数据",
    description="为测试目的生成随机合作数据",
)
async def generate_sample_collaborations(
    num_samples: int = Query(100, ge=10, le=1000, description="生成合作数量"),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Generate sample collaboration data for testing.
    """
    from app.services.collaboration_service import CollaborationService

    service = CollaborationService(session)
    try:
        count = await service.generate_sample_collaborations(num_samples)
        return {"message": f"已生成 {count} 条合作数据", "count": count}
    finally:
        await service.close()


@router.get(
    "/{talent_id}/collaborations",
    summary="获取合作网络",
    description="获取学者的合作关系数据",
)
async def get_talent_collaborations(
    talent_id: int,
    limit: int = Query(20, ge=1, le=50, description="返回合作者数量限制"),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Get collaboration network for a talent.
    """
    from app.services.collaboration_service import CollaborationService

    talent_service = TalentService(session)

    # Verify talent exists
    if not await talent_service.talent_exists(talent_id):
        raise HTTPException(status_code=404, detail="Talent not found")

    # Get collaboration network
    collab_service = CollaborationService(session)
    try:
        network = await collab_service.get_collaboration_network(talent_id, limit)
        return network
    finally:
        await collab_service.close()


# Global state for tracking sync progress
_sync_progress = {"status": "idle", "processed": 0, "total": 0, "collaborations": 0}


@router.post(
    "/collaborations/sync",
    summary="同步合作网络数据",
    description="从已采集的论文数据中提取学者合作关系，无需重复调用 OpenAlex API",
)
async def sync_collaborations(
    background_tasks: BackgroundTasks,
    talent_id: int | None = Query(None, description="单个学者ID，为空则同步全部"),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Trigger collaboration data sync from local RawWork data.
    """
    global _sync_progress

    if _sync_progress["status"] == "running":
        raise HTTPException(status_code=409, detail="同步任务正在进行中，请稍后再试")

    # Reset progress
    _sync_progress = {"status": "pending", "processed": 0, "total": 0, "collaborations": 0}

    # Use FastAPI BackgroundTasks instead of manual threading
    # This keeps the task in the same event loop, avoiding "Future attached to a different loop" errors
    background_tasks.add_task(run_sync_background, talent_id)

    return {
        "message": "同步任务已启动",
        "talent_id": talent_id,
        "sync_all": talent_id is None
    }


async def run_sync_background(talent_id: int | None = None):
    """Run sync as a background task within the main event loop."""
    global _sync_progress
    from app.core.database import AsyncSessionLocal
    from app.services.collaboration_service import CollaborationService

    _sync_progress = {"status": "running", "processed": 0, "total": 0, "collaborations": 0}

    async with AsyncSessionLocal() as session:
        service = CollaborationService(session)
        try:
            if talent_id:
                from app.services.talent_service import TalentService
                talent_service = TalentService(session)
                talent = await talent_service.get_talent_by_id(talent_id)
                if talent:
                    count = await service.sync_collaborations_for_talent(talent)
                    _sync_progress["processed"] = 1
                    _sync_progress["total"] = 1
                    _sync_progress["collaborations"] = count
            else:
                result = await service.sync_all_collaborations(
                    progress_callback=lambda p, t, c: _sync_progress.update({
                        "processed": p, "total": t, "collaborations": c
                    })
                )
                _sync_progress.update(result)

            _sync_progress["status"] = "completed"
        except Exception as e:
            import traceback
            traceback.print_exc()
            _sync_progress["status"] = f"error: {str(e)}"


@router.get(
    "/collaborations/status",
    summary="获取同步状态",
    description="获取合作网络数据同步的进度状态",
)
async def get_sync_status(
    session: AsyncSession = Depends(get_async_session),
):
    """
    Get collaboration sync status.
    """
    global _sync_progress
    from app.services.collaboration_service import CollaborationService

    # Get current data status
    service = CollaborationService(session)
    try:
        data_status = await service.get_sync_status()
        return {
            "sync_progress": _sync_progress,
            "data_status": data_status
        }
    finally:
        await service.close()
