"""
Talent export/compare endpoints.
人才导出与对比接口

Split from talents.py; routes keep the original /talents prefix.

Architecture: Endpoint -> Service -> Repository
"""

from __future__ import annotations

import csv
import io

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_async_session
from app.domains.academic.schemas.overview import TalentCompareResponse
from app.domains.academic.services.talent_service import TalentService
from app.domains.shared.api.auth import require_admin
from app.domains.shared.services.audit_service import AuditService

router = APIRouter(prefix="/talents", tags=["Talents"])


@router.post(
    "/export",
    response_model=None,
    summary="导出候选人",
    description="导出选中的候选人数据为CSV或Excel格式",
)
async def export_talents(
    format: str = Query("csv", description="导出格式: csv 或 xlsx"),
    talent_ids: list[int] = Query(..., description="要导出的人才ID列表"),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(require_admin),
    request: Request = None,
):
    """
    Export selected talents to CSV or Excel format.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    service = TalentService(session)

    # Fetch talents by IDs
    talents = await service.get_talents_by_ids(talent_ids)

    if not talents:
        await AuditService.log_data_operation(
            user_id=current_user.get("user_id"),
            operation="export",
            resource_type="talent",
            resource_id=None,
            status="failure",
            user_ip=request.client.host if request and request.client else None,
            request_id=getattr(request.state, "request_id", None) if request else None,
            detail={"format": format, "talent_ids": talent_ids, "error": "未找到要导出的人才"},
        )
        raise HTTPException(status_code=404, detail="未找到要导出的人才")

    # Watermark disclaimer text
    disclaimer = (
        "【重要声明】本文件导出的人才数据仅供内部人才发现与学术调研使用。"
        "严禁通过任何渠道向人才发起招聘邀约，严禁将数据提供给第三方招聘机构，"
        "严禁用于商业营销或数据贩卖。违规使用将导致账号封禁及法律责任。"
    )

    # Prepare data
    headers = [
        "ID",
        "姓名",
        "英文名",
        "ORCID",
        "角色",
        "学校",
        "职位",
        "论文数",
        "引用数",
        "H指数",
        "研究方向",
    ]
    rows = []
    for t in talents:
        rows.append(
            [
                t.talent_id,
                t.name,
                t.name_en or "",
                t.orcid or "",
                t.role_type,
                t.school.school_name if t.school else "",
                t.current_title or "",
                t.works_count,
                t.cited_by_count,
                t.h_index,
                ", ".join(t.topic_tags or []),
            ]
        )

    if format == "xlsx":
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "候选人导出"

        # Write disclaimer in first row, spanning all columns
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        disclaimer_cell = ws.cell(row=1, column=1, value=disclaimer)
        disclaimer_cell.alignment = Alignment(wrap_text=True, vertical="center")
        disclaimer_cell.font = Font(color="FF0000", bold=True)
        ws.row_dimensions[1].height = 45

        # Write headers (row 3, leaving row 2 as blank separator)
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)

        # Write data
        for row_idx, row in enumerate(rows, 4):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        await AuditService.log_data_operation(
            user_id=current_user.get("user_id"),
            operation="export",
            resource_type="talent",
            resource_id=None,
            status="success",
            user_ip=request.client.host if request and request.client else None,
            request_id=getattr(request.state, "request_id", None) if request else None,
            detail={"format": format, "count": len(talents), "talent_ids": talent_ids},
        )

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=talents_export.xlsx"},
        )
    else:
        # Create CSV
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([disclaimer])
        writer.writerow([])
        writer.writerow(headers)
        writer.writerows(rows)
        buffer.seek(0)

        response = StreamingResponse(
            io.BytesIO(buffer.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=talents_export.csv"},
        )

        await AuditService.log_data_operation(
            user_id=current_user.get("user_id"),
            operation="export",
            resource_type="talent",
            resource_id=None,
            status="success",
            user_ip=request.client.host if request and request.client else None,
            request_id=getattr(request.state, "request_id", None) if request else None,
            detail={"format": format, "count": len(talents), "talent_ids": talent_ids},
        )

        return response


@router.post(
    "/compare",
    response_model=TalentCompareResponse,
    summary="对比候选人",
    description="获取多个候选人的对比数据",
)
async def compare_talents(
    talent_ids: list[int] = Query(..., description="要对比的人才ID列表 (2-4个)"),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Compare multiple talents (2-4) side by side.
    Returns detailed comparison data.
    """
    if len(talent_ids) < 2 or len(talent_ids) > 4:
        raise HTTPException(status_code=400, detail="请选择2-4位候选人进行对比")

    service = TalentService(session)

    # Fetch talents by IDs
    talents = await service.get_talents_by_ids(talent_ids)

    if not talents:
        raise HTTPException(status_code=404, detail="未找到要对比的人才")

    # Build comparison data
    from app.domains.academic.schemas.overview import ComparisonFieldItem, TalentCompareItem

    comparison_data = []
    for t in talents:
        comparison_data.append(
            TalentCompareItem(
                talent_id=t.talent_id,
                name=t.name,
                name_en=t.name_en,
                orcid=t.orcid,
                role_type=t.role_type,
                school_id=t.school_id,
                school_name=t.primary_school_name,
                current_title=t.current_title,
                department_name=t.department_name,
                lab_name=t.lab_name,
                works_count=t.works_count,
                cited_by_count=t.cited_by_count,
                h_index=t.h_index,
                latest_active_year=t.latest_active_year,
                topic_tags=t.topic_tags or [],
                openalex_topics=t.openalex_topics or [],
                academic_age=t.role_profile.academic_age if t.role_profile else None,
            )
        )

    return TalentCompareResponse(
        talents=comparison_data,
        comparison_fields=[
            ComparisonFieldItem(key="name", label="姓名"),
            ComparisonFieldItem(key="role_type", label="角色"),
            ComparisonFieldItem(key="school_name", label="学校"),
            ComparisonFieldItem(key="current_title", label="职位"),
            ComparisonFieldItem(key="department_name", label="院系"),
            ComparisonFieldItem(key="works_count", label="论文数"),
            ComparisonFieldItem(key="cited_by_count", label="引用数"),
            ComparisonFieldItem(key="h_index", label="H指数"),
            ComparisonFieldItem(key="latest_active_year", label="最近活跃年份"),
            ComparisonFieldItem(key="academic_age", label="学术年龄"),
            ComparisonFieldItem(key="topic_tags", label="研究方向"),
        ],
    )
