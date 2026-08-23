"""OS developer export - Excel/CSV assembly for the export endpoint.

Extracted from ``api/developers.py::export_developers`` (2026-08 cohesion
refactor): column definitions, disclaimer text, label mapping, file building
and audit logging live here; the endpoint shrinks to parameter parsing and
delegation. Behavior locked by tests/domains/open_source/test_developer_export.py.
"""

from __future__ import annotations

import csv
import io
from typing import Any, cast
from urllib.parse import quote

from fastapi import Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.core.exceptions import NotFoundError
from app.domains.open_source.models.open_source import OSDeveloper
from app.domains.open_source.services.open_source_service import OpenSourceService
from app.domains.shared.constants.tech_taxonomy import TECH_DOMAINS, TECH_ELEMENTS
from app.domains.shared.services.audit_service import AuditService

_DOMAIN_NAMES = {d["code"]: d["name"] for d in TECH_DOMAINS}

DISCLAIMER = (
    "【重要声明】本文件导出的开源人才数据仅供内部人才发现与学术调研使用。"
    "严禁通过任何渠道向人才发起招聘邀约，严禁将数据提供给第三方招聘机构，"
    "严禁用于商业营销或数据贩卖。违规使用将导致账号封禁及法律责任。"
)

EXPORT_HEADERS = [
    "序号",
    "GitHub账号",
    "GitHub主页",
    "姓名",
    "邮箱",
    "公司",
    "地区",
    "Blog主页",
    "Stars总数",
    "仓库数",
    "Followers数",
    "主要语言",
    "技术领域",
    "技术要素/方向",
    "收录来源的开源项目",
    "社交媒体链接（供参考）",
]


def _tech_labels(tech_tags: list[str] | None) -> tuple[str, str]:
    """Map element codes to (领域中文, 要素中文) — domain labels deduped."""
    domains: list[str] = []
    elements: list[str] = []
    for code in tech_tags or []:
        el = TECH_ELEMENTS.get(code)
        if not el:
            elements.append(code)
            continue
        domain_name = _DOMAIN_NAMES.get(el["domain"], el["domain"])
        if domain_name not in domains:
            domains.append(domain_name)
        if el["name"] not in elements:
            elements.append(el["name"])
    return "、".join(domains), "、".join(elements)


class OSDeveloperExporter:
    """Build the developer export file (CSV/XLSX) and audit the operation."""

    def __init__(self, service: OpenSourceService) -> None:
        self.service = service

    async def export(
        self,
        developer_ids: list[int],
        fmt: str,
        current_user: dict,
        request: Request | None,
    ) -> StreamingResponse:
        developers = await self.service.get_developers_by_ids(developer_ids)

        if not developers:
            await self._log_operation(
                status="failure",
                current_user=current_user,
                request=request,
                detail={
                    "format": fmt,
                    "developer_ids": developer_ids,
                    "error": "未找到要导出的开发者",
                },
            )
            raise NotFoundError("未找到要导出的开发者")

        # Batch fetch collected repos for enriched fields
        dev_ids = [cast(int, d.developer_id) for d in developers]
        repos_map = await self.service.get_collected_repos_for_developers(dev_ids)

        rows = self._build_rows(developers, repos_map)

        if fmt == "xlsx":
            response = self._build_xlsx(rows)
        else:
            response = self._build_csv(rows)

        await self._log_operation(
            status="success",
            current_user=current_user,
            request=request,
            detail={
                "format": fmt,
                "count": len(developers),
                "developer_ids": developer_ids,
            },
        )
        return response

    @staticmethod
    async def _log_operation(
        status: str,
        current_user: dict,
        request: Request | None,
        detail: dict[str, Any],
    ) -> None:
        await AuditService.log_data_operation(
            user_id=current_user.get("user_id"),
            operation="export",
            resource_type="os_developer",
            resource_id=None,
            status=status,
            user_ip=request.client.host if request and request.client else None,
            request_id=getattr(request.state, "request_id", None) if request else None,
            detail=detail,
        )

    @staticmethod
    def _build_rows(
        developers: list[OSDeveloper],
        repos_map: dict[int, list[str]],
    ) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for idx, d in enumerate(developers, 1):
            repo_names = ", ".join(repos_map.get(cast(int, d.developer_id), []))
            name = d.name or d.github_login or ""
            company = d.company or ""
            search_query = f"{name} {company} LinkedIn".strip()
            socials = cast("dict[str, str] | None", d.social_links) or {}
            if socials:
                # Real profile links first (twitter/linkedin/website…), one per platform
                social_link = ", ".join(socials.values())
            else:
                social_link = f"https://www.google.com/search?q={quote(search_query, safe='')}"
            tech_domains, tech_elements = _tech_labels(cast("list[str] | None", d.tech_tags))

            rows.append(
                [
                    idx,
                    d.github_login,
                    f"https://github.com/{d.github_login}",
                    d.name or "",
                    d.email or "",
                    d.company or "",
                    d.location or "",
                    d.blog_url or "",
                    d.total_stars_received,
                    d.public_repos_count,
                    d.followers_count,
                    ", ".join(d.primary_languages or []),
                    tech_domains,
                    tech_elements,
                    repo_names,
                    social_link,
                ]
            )
        return rows

    @staticmethod
    def _build_xlsx(rows: list[list[Any]]) -> StreamingResponse:
        wb = Workbook()
        ws = wb.active
        ws.title = "开源人才导出"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EXPORT_HEADERS))
        disclaimer_cell = ws.cell(row=1, column=1, value=DISCLAIMER)
        disclaimer_cell.alignment = Alignment(wrap_text=True, vertical="center")
        disclaimer_cell.font = Font(color="FF0000", bold=True)
        ws.row_dimensions[1].height = 45

        for col, header in enumerate(EXPORT_HEADERS, 1):
            ws.cell(row=3, column=col, value=header)

        for row_idx, row in enumerate(rows, 4):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for idx in range(1, len(EXPORT_HEADERS) + 1):
            max_length = 0
            column_letter = get_column_letter(idx)
            # Skip disclaimer row (row 1) and empty separator row (row 2);
            # start from header row (row 3)
            for cell in ws[column_letter][2:]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=os_developers_export.xlsx"},
        )

    @staticmethod
    def _build_csv(rows: list[list[Any]]) -> StreamingResponse:
        text_buffer = io.StringIO()
        writer = csv.writer(text_buffer)
        writer.writerow([DISCLAIMER])
        writer.writerow([])
        writer.writerow(EXPORT_HEADERS)
        writer.writerows(rows)
        text_buffer.seek(0)

        return StreamingResponse(
            io.BytesIO(text_buffer.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=os_developers_export.csv"},
        )
