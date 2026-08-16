"""
Open Source — Developer, Repository, and Search endpoints.
"""

from __future__ import annotations

from typing import cast

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_async_session
from app.domains.open_source.api.auth import get_current_user
from app.domains.open_source.schemas.open_source import (
    OSContributionItem,
    OSDeveloperCompareRequest,
    OSDeveloperCompareResponse,
    OSDeveloperDetail,
    OSDeveloperSummary,
    OSLanguageSkillItem,
    OSRepoConfigResponse,
    OSRepositoryContributor,
    OSRepositoryDetailResponse,
    OSRepositoryItem,
    OSSearchRequest,
)
from app.domains.open_source.services.open_source_service import OpenSourceService
from app.domains.shared.api.auth import require_admin
from app.domains.shared.constants.tech_taxonomy import TECH_DOMAINS, TECH_ELEMENTS
from app.domains.shared.schemas.common import PaginatedResponse
from app.domains.shared.services.audit_service import AuditService

router = APIRouter(prefix="/open-source", tags=["Open Source Talent"])

_DOMAIN_NAMES = {d["code"]: d["name"] for d in TECH_DOMAINS}


def _tech_labels(tech_tags: list[str] | None) -> tuple[str, str]:
    """Map element codes to (领域中文, 要素中文) — domain labels deduped."""
    domains: list[str] = []
    elements: list[str] = []
    for code in tech_tags or []:
        el = TECH_ELEMENTS.get(code)
        if not el:
            elements.append(code)
            continue
        domain_name = _DOMAIN_NAMES.get(el["domain"], el["domain"])
        if domain_name not in domains:
            domains.append(domain_name)
        if el["name"] not in elements:
            elements.append(el["name"])
    return "、".join(domains), "、".join(elements)


# ============= Developers =============


@router.get("/developers", response_model=PaginatedResponse[OSDeveloperSummary])
async def list_developers(
    q: str = Query("", description="Keyword search"),
    tech_elements: list[str] | None = Query(None),
    languages: list[str] | None = Query(None),
    location: str | None = Query(None),
    company: str | None = Query(None),
    min_stars: int | None = Query(None, ge=0),
    is_committer: bool | None = Query(None, description="Filter developers who are committers"),
    is_student: bool | None = Query(None, description="Filter developers who are students"),
    repo_full_names: list[str] | None = Query(None, description="Filter by repository full names"),
    sort_by: str = Query("stars_desc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
) -> PaginatedResponse[OSDeveloperSummary]:
    service = OpenSourceService(session)
    items, total = await service.list_developers(
        q=q,
        tech_elements=tech_elements,
        languages=languages,
        location=location,
        company=company,
        min_stars=min_stars,
        is_committer=is_committer,
        is_student=is_student,
        repo_full_names=repo_full_names,
        sort_by=sort_by,
        page=page,
        page_size=page_size,
    )

    # Enrich summaries with aggregated role tags from contributions (single batch query)
    dev_ids = [cast(int, dev.developer_id) for dev in items]
    roles_map = await service.get_developer_roles_map(dev_ids)
    summaries: list[OSDeveloperSummary] = []
    for dev, dev_id in zip(items, dev_ids, strict=True):
        summary = OSDeveloperSummary.model_validate(dev)
        summary.roles = roles_map.get(dev_id, [])
        summaries.append(summary)

    return PaginatedResponse.create(
        items=summaries,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/developers/ids", response_model=list[int])
async def list_all_developer_ids(
    q: str = Query("", description="Keyword search"),
    tech_elements: list[str] | None = Query(None),
    languages: list[str] | None = Query(None),
    location: str | None = Query(None),
    company: str | None = Query(None),
    min_stars: int | None = Query(None, ge=0),
    is_committer: bool | None = Query(None, description="Filter developers who are committers"),
    is_student: bool | None = Query(None, description="Filter developers who are students"),
    repo_full_names: list[str] | None = Query(None, description="Filter by repository full names"),
    sort_by: str = Query("stars_desc"),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
) -> list[int]:
    """Return all developer IDs matching the current filters (no pagination).
    Used for frontend 'select all' feature.
    """
    service = OpenSourceService(session)
    items, _total = await service.list_developers(
        q=q,
        tech_elements=tech_elements,
        languages=languages,
        location=location,
        company=company,
        min_stars=min_stars,
        is_committer=is_committer,
        is_student=is_student,
        repo_full_names=repo_full_names,
        sort_by=sort_by,
        page=1,
        page_size=100000,
    )
    return [cast(int, dev.developer_id) for dev in items]


@router.get("/developers/{developer_id}", response_model=OSDeveloperDetail)
async def get_developer(
    developer_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
) -> OSDeveloperDetail:
    service = OpenSourceService(session)
    return await service.get_developer_detail(developer_id)


@router.get(
    "/developers/{developer_id}/repositories", response_model=PaginatedResponse[OSRepositoryItem]
)
async def list_developer_repositories(
    developer_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_by: str = Query("stars"),
    session: AsyncSession = Depends(get_async_session),
) -> PaginatedResponse[OSRepositoryItem]:
    service = OpenSourceService(session)
    items = await service.get_developer_repositories(developer_id)
    # Simple in-memory sort/paginate
    reverse = sort_by != "name"
    items = sorted(
        items,
        key=lambda r: getattr(
            r,
            (
                (sort_by.replace("_desc", "").replace("_asc", "") + "_count")
                if sort_by in ("stars", "forks")
                else "name"
            ),
        ),
        reverse=reverse,
    )
    total = len(items)
    start = (page - 1) * page_size
    items = items[start : start + page_size]
    return PaginatedResponse.create(
        items=[OSRepositoryItem.model_validate(i) for i in items],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/developers/{developer_id}/contributions", response_model=list[OSContributionItem])
async def list_developer_contributions(
    developer_id: int,
    session: AsyncSession = Depends(get_async_session),
) -> list[OSContributionItem]:
    service = OpenSourceService(session)
    result = await service.get_developer_contributions(developer_id)
    return [
        OSContributionItem(
            contribution_id=cast(int, c.contribution_id),
            repo_id=cast(int, c.repo_id),
            repo_full_name=full_name,
            commits_count=cast(int, c.commits_count),
            prs_count=cast(int, c.prs_count),
            issues_count=cast(int, c.issues_count),
            code_reviews_count=cast(int, c.code_reviews_count),
            is_owner=cast(bool, c.is_owner),
            is_maintainer=cast(bool, c.is_maintainer),
            is_committer=cast(bool, c.is_committer),
        )
        for c, full_name in result
    ]


@router.get("/developers/{developer_id}/languages", response_model=list[OSLanguageSkillItem])
async def list_developer_languages(
    developer_id: int,
    session: AsyncSession = Depends(get_async_session),
) -> list[OSLanguageSkillItem]:
    service = OpenSourceService(session)
    items = await service.get_developer_languages(developer_id)
    return [OSLanguageSkillItem.model_validate(i) for i in items]


@router.post("/developers/compare", response_model=OSDeveloperCompareRequest)
async def compare_developers(
    data: OSDeveloperCompareRequest,
    session: AsyncSession = Depends(get_async_session),
) -> OSDeveloperCompareResponse:
    service = OpenSourceService(session)
    return await service.compare_developers(data.developer_ids)


@router.get("/developers/{developer_id}/recommend", response_model=list[OSDeveloperSummary])
async def recommend_similar_developers(
    developer_id: int,
    limit: int = Query(10, ge=1, le=50),
    session: AsyncSession = Depends(get_async_session),
) -> list[OSDeveloperSummary]:
    service = OpenSourceService(session)
    items = await service.recommend_similar(developer_id, limit=limit)
    return [OSDeveloperSummary.model_validate(i) for i in items]


class ExportDevelopersRequest(BaseModel):
    developer_ids: list[int]
    format: str = "csv"


@router.post("/developers/export", response_model=None)
async def export_developers(
    data: ExportDevelopersRequest,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(require_admin),
    request: Request = None,  # type: ignore[assignment]
) -> StreamingResponse:
    """Export selected developers to CSV or Excel format."""
    import csv
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    service = OpenSourceService(session)
    developers = await service.get_developers_by_ids(data.developer_ids)

    if not developers:
        await AuditService.log_data_operation(
            user_id=current_user.get("user_id"),
            operation="export",
            resource_type="os_developer",
            resource_id=None,
            status="failure",
            user_ip=request.client.host if request and request.client else None,
            request_id=getattr(request.state, "request_id", None) if request else None,
            detail={
                "format": data.format,
                "developer_ids": data.developer_ids,
                "error": "未找到要导出的开发者",
            },
        )
        raise HTTPException(status_code=404, detail="未找到要导出的开发者")

    # Batch fetch collected repos for enriched fields
    dev_ids = [cast(int, d.developer_id) for d in developers]
    repos_map = await service.get_collected_repos_for_developers(dev_ids)

    disclaimer = (
        "【重要声明】本文件导出的开源人才数据仅供内部人才发现与学术调研使用。"
        "严禁通过任何渠道向人才发起招聘邀约，严禁将数据提供给第三方招聘机构，"
        "严禁用于商业营销或数据贩卖。违规使用将导致账号封禁及法律责任。"
    )

    headers = [
        "序号",
        "GitHub账号",
        "GitHub主页",
        "姓名",
        "邮箱",
        "公司",
        "地区",
        "Blog主页",
        "Stars总数",
        "仓库数",
        "Followers数",
        "主要语言",
        "技术领域",
        "技术要素/方向",
        "收录来源的开源项目",
        "社交媒体链接（供参考）",
    ]
    from urllib.parse import quote

    rows = []
    for idx, d in enumerate(developers, 1):
        repo_names = ", ".join(repos_map.get(cast(int, d.developer_id), []))
        name = d.name or d.github_login or ""
        company = d.company or ""
        search_query = f"{name} {company} LinkedIn".strip()
        social_link = f"https://www.google.com/search?q={quote(search_query, safe='')}"
        tech_domains, tech_elements = _tech_labels(d.tech_tags)

        rows.append(
            [
                idx,
                d.github_login,
                f"https://github.com/{d.github_login}",
                d.name or "",
                d.email or "",
                d.company or "",
                d.location or "",
                d.blog_url or "",
                d.total_stars_received,
                d.public_repos_count,
                d.followers_count,
                ", ".join(d.primary_languages or []),
                tech_domains,
                tech_elements,
                repo_names,
                social_link,
            ]
        )

    if data.format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "开源人才导出"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        disclaimer_cell = ws.cell(row=1, column=1, value=disclaimer)
        disclaimer_cell.alignment = Alignment(wrap_text=True, vertical="center")
        disclaimer_cell.font = Font(color="FF0000", bold=True)
        ws.row_dimensions[1].height = 45

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)

        for row_idx, row in enumerate(rows, 4):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        from openpyxl.utils import get_column_letter

        for idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(idx)
            # Skip disclaimer row (row 1) and empty separator row (row 2);
            # start from header row (row 3)
            for cell in ws[column_letter][2:]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        await AuditService.log_data_operation(
            user_id=current_user.get("user_id"),
            operation="export",
            resource_type="os_developer",
            resource_id=None,
            status="success",
            user_ip=request.client.host if request and request.client else None,
            request_id=getattr(request.state, "request_id", None) if request else None,
            detail={
                "format": data.format,
                "count": len(developers),
                "developer_ids": data.developer_ids,
            },
        )

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=os_developers_export.xlsx"},
        )
    else:
        text_buffer = io.StringIO()
        writer = csv.writer(text_buffer)
        writer.writerow([disclaimer])
        writer.writerow([])
        writer.writerow(headers)
        writer.writerows(rows)
        text_buffer.seek(0)

        await AuditService.log_data_operation(
            user_id=current_user.get("user_id"),
            operation="export",
            resource_type="os_developer",
            resource_id=None,
            status="success",
            user_ip=request.client.host if request and request.client else None,
            request_id=getattr(request.state, "request_id", None) if request else None,
            detail={
                "format": data.format,
                "count": len(developers),
                "developer_ids": data.developer_ids,
            },
        )

        return StreamingResponse(
            io.BytesIO(text_buffer.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=os_developers_export.csv"},
        )


# ============= Public Repository List =============


@router.get("/repositories", response_model=PaginatedResponse[OSRepoConfigResponse])
async def list_public_repositories(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    tech_elements: list[str] | None = Query(None, description="Filter by tech elements"),
    q: str | None = Query(None, description="Search by repo name or description"),
    sort_by: str = Query("stars", description="stars | id_desc"),
    collected_only: bool = Query(True, description="Only repos with completed collect tasks"),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
) -> PaginatedResponse[OSRepoConfigResponse]:
    """List all public repositories (collected repo configs)."""
    service = OpenSourceService(session)
    items, total = await service.list_repo_configs(
        page=page,
        page_size=page_size,
        tech_elements=tech_elements,
        is_active=True,
        collect_enabled=None,
        sort_by=sort_by,
        collected_only=collected_only,
        q=q,
    )
    return PaginatedResponse.create(
        items=[OSRepoConfigResponse.model_validate(i) for i in items],
        total=total,
        page=page,
        page_size=page_size,
    )


# ============= Repository (Project) Detail =============


@router.get("/repositories/{owner}/{name}", response_model=OSRepositoryDetailResponse)
async def get_repository(
    owner: str,
    name: str,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
) -> OSRepositoryDetailResponse:
    """Get repository detail with contributor count by full name."""
    service = OpenSourceService(session)
    return cast(OSRepositoryDetailResponse, await service.get_repository_detail(f"{owner}/{name}"))


@router.get(
    "/repositories/{owner}/{name}/contributors",
    response_model=PaginatedResponse[OSRepositoryContributor],
)
async def get_repository_contributors(
    owner: str,
    name: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
) -> PaginatedResponse[OSRepositoryContributor]:
    """Get contributors for a repository by full name."""
    service = OpenSourceService(session)
    items, total = await service.get_repository_contributors(f"{owner}/{name}", page, page_size)
    return PaginatedResponse.create(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
    )


# ============= Search (v2 unified) =============


@router.post("/search", response_model=PaginatedResponse[OSDeveloperSummary])
async def search_developers(
    req: OSSearchRequest,
    session: AsyncSession = Depends(get_async_session),
    current_user: dict = Depends(get_current_user),
) -> PaginatedResponse[OSDeveloperSummary]:
    """Unified search endpoint supporting keyword/semantic/hybrid modes."""
    service = OpenSourceService(session)
    items, total = await service.search_developers(req)
    return PaginatedResponse.create(
        items=[OSDeveloperSummary.model_validate(i) for i in items],
        total=total,
        page=req.page,
        page_size=req.page_size,
    )
