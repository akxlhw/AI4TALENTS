"""Characterization tests for the open-source developer export endpoint.

Locks the export contract (column headers, disclaimer text, CSV/XLSX shapes,
admin-only access, empty-selection 404) BEFORE the assembly logic moves out of
the endpoint into a dedicated exporter service.
"""

from __future__ import annotations

import csv
import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.domains.open_source.models.open_source import OSDeveloper
from app.main import app

DISCLAIMER_PREFIX = "【重要声明】"

EXPECTED_HEADERS = [
    "序号",
    "GitHub账号",
    "GitHub主页",
    "姓名",
    "邮箱",
    "公司",
    "地区",
    "Blog主页",
    "Stars总数",
    "仓库数",
    "Followers数",
    "主要语言",
    "技术领域",
    "技术要素/方向",
    "收录来源的开源项目",
    "社交媒体链接（供参考）",
]


@pytest.fixture
async def export_developers(test_session: AsyncSession) -> list[OSDeveloper]:
    """Two developers: one rich profile (google-search fallback link),
    one with a real social link."""
    devs = [
        OSDeveloper(
            github_login="alice",
            name="Alice Zhang",
            company="Meta",
            location="SF",
            email="alice@example.com",
            blog_url="https://alice.dev",
            followers_count=100,
            public_repos_count=10,
            total_stars_received=500,
            primary_languages=["Python", "Go"],
        ),
        OSDeveloper(
            github_login="bob",
            name="Bob Li",
            social_links={"linkedin": "https://linkedin.com/in/bob"},
        ),
    ]
    test_session.add_all(devs)
    await test_session.commit()
    return devs


@pytest.fixture
async def admin_client(client: AsyncClient) -> AsyncClient:
    """Bypass require_admin with a super_admin identity."""
    from app.domains.shared.api.auth import require_admin

    app.dependency_overrides[require_admin] = lambda: {
        "user_id": 1,
        "username": "export_admin",
        "role": "super_admin",
    }
    yield client
    app.dependency_overrides.pop(require_admin, None)


@pytest.fixture
async def user_client(client: AsyncClient) -> AsyncClient:
    """Authenticated NON-admin identity: only require_user is stubbed so the
    real require_admin role check runs and must reject with 403."""
    from app.domains.shared.api.auth_deps import require_user

    app.dependency_overrides[require_user] = lambda: {
        "user_id": 2,
        "username": "normal_user",
        "role": "user",
    }
    yield client
    app.dependency_overrides.pop(require_user, None)


def _export_payload(devs: list[OSDeveloper], fmt: str) -> dict:
    return {
        "developer_ids": [d.developer_id for d in devs],
        "format": fmt,
    }


# ============ CSV contract ============


@pytest.mark.asyncio
async def test_csv_export_contract(
    admin_client: AsyncClient, export_developers: list[OSDeveloper]
) -> None:
    resp = await admin_client.post(
        "/api/v1/open-source/developers/export",
        json=_export_payload(export_developers, "csv"),
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    assert "os_developers_export.csv" in resp.headers["content-disposition"]

    # utf-8-sig BOM so Excel opens Chinese text correctly
    assert resp.content[:3] == b"\xef\xbb\xbf"
    text = resp.content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))

    # Row layout: disclaimer / blank separator / headers / data rows
    assert len(rows) == 3 + len(export_developers)
    assert len(rows[0]) == 1
    assert rows[0][0].startswith(DISCLAIMER_PREFIX)
    assert "严禁" in rows[0][0]
    assert rows[1] == []
    assert rows[2] == EXPECTED_HEADERS

    alice = rows[3]
    assert alice[0] == "1"
    assert alice[1] == "alice"
    assert alice[2] == "https://github.com/alice"
    assert alice[3] == "Alice Zhang"
    assert alice[4] == "alice@example.com"
    assert alice[5] == "Meta"
    assert alice[6] == "SF"
    assert alice[7] == "https://alice.dev"
    assert alice[8] == "500"
    assert alice[9] == "10"
    assert alice[10] == "100"
    assert alice[11] == "Python, Go"
    # No social links -> google search fallback
    assert alice[15].startswith("https://www.google.com/search?q=")

    bob = rows[4]
    assert bob[0] == "2"
    assert bob[1] == "bob"
    # Real social links are exported verbatim
    assert bob[15] == "https://linkedin.com/in/bob"


# ============ XLSX contract ============


@pytest.mark.asyncio
async def test_xlsx_export_contract(
    admin_client: AsyncClient, export_developers: list[OSDeveloper]
) -> None:
    resp = await admin_client.post(
        "/api/v1/open-source/developers/export",
        json=_export_payload(export_developers, "xlsx"),
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "os_developers_export.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.title == "开源人才导出"

    # Row 1 merged disclaimer, red bold font
    assert ws.cell(row=1, column=1).value.startswith(DISCLAIMER_PREFIX)
    # Header row at 3 (row 2 is the blank separator)
    headers = [ws.cell(row=3, column=c).value for c in range(1, len(EXPECTED_HEADERS) + 1)]
    assert headers == EXPECTED_HEADERS

    # Data rows follow
    assert ws.cell(row=4, column=1).value == 1
    assert ws.cell(row=4, column=2).value == "alice"
    assert ws.cell(row=5, column=2).value == "bob"
    assert ws.max_row == 3 + len(export_developers)


# ============ Access control & empty selection ============


@pytest.mark.asyncio
async def test_export_requires_admin(user_client: AsyncClient) -> None:
    resp = await user_client.post(
        "/api/v1/open-source/developers/export",
        json={"developer_ids": [1], "format": "csv"},
    )
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_export_unknown_ids_returns_404(admin_client: AsyncClient) -> None:
    resp = await admin_client.post(
        "/api/v1/open-source/developers/export",
        json={"developer_ids": [987654321], "format": "csv"},
    )
    assert resp.status_code == 404
